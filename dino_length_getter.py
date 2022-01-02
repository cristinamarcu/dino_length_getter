import logging
import re
import string
from typing import Optional
import bs4
import requests
import sys
import openpyxl


def get_dino_name(linkHref) -> Optional[str]:
    relinkdino = re.compile(r'/discover/dino-directory/([a-z]+).html')
    linkdino = relinkdino.findall(linkHref)
    if len(linkdino) < 1:
        return None
    logging.info(f'{linkdino} is the name of the dinosaur from {linkHref}')
    return linkdino[0]


def dinoname_by_letter(letter: str) -> list[str]:
    res = requests.get(r'https://www.nhm.ac.uk/discover/dino-directory/name/' + letter + '/gallery.html')
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    links = soup.findAll('a')
    dinonames = []
    for link in links:
        matchedDinoName = get_dino_name(link['href'])
        if matchedDinoName is not None:
            dinonames.append(matchedDinoName)
            logging.info(f'The dinosaur {matchedDinoName} was added to the list.')
    return dinonames


def getlength(dinoName) -> Optional[str]:
    res = requests.get(r'https://www.nhm.ac.uk/discover/dino-directory/' + dinoName + '.html')

    if res.status_code != 200:
        logging.error(f'The dinosaur {dinoName} is not in dino-directory.')
        return None

    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    length = soup.select(
        '#main-content > div > div > div.row2cells13.section > div > div.large-9.medium-9.columns > div > div > div > div:nth-child(3) > div.dinosaur--description-container.small-12.medium-12.large-5.columns > dl > dd:nth-child(4)')
    if len(length) == 0:
        logging.error(f'{dinoName} does not have a length on the website.')
        return 'not available'
    logging.info(f'{dinoName} has the length {length[0].text}.')
    return length[0].text


def dinoWorksheet(wsheet):
    nrdino = 0
    id0 = 1
    for letter in string.ascii_lowercase:
        dinoname_by_letter_list = dinoname_by_letter(letter)
        nrdino = nrdino + len(dinoname_by_letter_list)
        j = 0
        i = id0
        while i < id0 + nrdino and j < len(dinoname_by_letter_list):
            wsheet.cell(row=i, column=1).value = dinoname_by_letter_list[j]
            wsheet.cell(row=i, column=2).value = getlength(dinoname_by_letter_list[j])
            i = i + 1
            j = j + 1
        id0 = nrdino + 1


"""
This program gets dinosaur length info from www.nhm.ac.uk and saves it to an excel worksheet
Example usage:
dino_length_getter.py diplodocus C:\\dinofolder\\dinosheet.xslx
"""
logging.basicConfig(level=logging.INFO)
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet.title = 'dinosheet'
dinoWorksheet(sheet)
wb.save(sys.argv[1])
